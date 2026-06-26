"""
process_gene_expression.py
---------------------------
Core qPCR analysis logic — pure Python, no LibreOffice dependency.

Spreadsheet format (per organ block):
  Optional text header row: col A = organ name, col C = HK gene name, col D = test gene name
  Data rows: col B = sample label, col C = HK value (float), col D = test value (float)
  Blocks separated by blank rows (no numeric C+D).

Control vs drug grouping:
  First named group per block = control.
  Row whose col B is blank / starts with "cont" = replicate of current group.
  Every subsequent named group = drug / treatment.
"""

import io
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COL_ORGAN = 1   # A
COL_LABEL = 2   # B
COL_HK    = 3   # C
COL_TEST  = 4   # D
COL_DELTA = 5   # E
COL_REL   = 6   # F

CTRL_HEX = "D9D9D9"
DRUG_HEX = "F4A99A"
HDR_HEX  = "2F4F8F"
CTRL_MPL = "#C8C8C8"
DRUG_MPL = "#E04030"


def cl(n):
    return get_column_letter(n)

def is_numeric(v):
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False

def is_continuation(label):
    if label is None:
        return True
    s = str(label).strip().lower()
    return s == "" or s.startswith("cont")


# ── Parser ────────────────────────────────────────────────────────────────────

def parse_blocks(ws) -> list:
    """
    Return one dict per organ block:
    {
        organ, hk_name, test_name,
        header_row (the text/label row just above first data row, or None),
        first_data_row, last_data_row,
        ctrl_rows, ctrl_label,
        drug_rows, drug_labels,
    }
    """
    blocks     = []
    max_row    = ws.max_row

    # Identify every row that has numeric values in both C and D
    numeric_rows = set()
    for r in range(1, max_row + 1):
        if is_numeric(ws.cell(row=r, column=COL_HK).value) and \
           is_numeric(ws.cell(row=r, column=COL_TEST).value):
            numeric_rows.add(r)

    visited = set()
    r = 1
    while r <= max_row:
        if r not in numeric_rows or r in visited:
            r += 1
            continue

        # ── find the contiguous block of numeric rows starting at r ──
        block_start = r
        block_end   = r
        while block_end + 1 in numeric_rows:
            block_end += 1

        for rr in range(block_start, block_end + 1):
            visited.add(rr)

        # ── look for organ name / column headers in the row(s) above ──
        organ     = None
        hk_name   = None
        test_name = None

        for look in range(block_start - 1, max(0, block_start - 4), -1):
            a = ws.cell(row=look, column=COL_ORGAN).value
            c = ws.cell(row=look, column=COL_HK).value
            d = ws.cell(row=look, column=COL_TEST).value
            if a is not None and str(a).strip():
                organ = str(a).strip()
            if isinstance(c, str) and c.strip():
                hk_name = c.strip()
            if isinstance(d, str) and d.strip():
                test_name = d.strip()
            if organ:
                break

        # Also check col A within the data rows themselves
        if organ is None:
            for rr in range(block_start, block_end + 1):
                a = ws.cell(row=rr, column=COL_ORGAN).value
                if a is not None and str(a).strip():
                    organ = str(a).strip()
                    break

        if organ is None:
            organ = f"Block_{block_start}"

        # ── group data rows by sample ──
        groups      = []
        current_grp = None

        for rr in range(block_start, block_end + 1):
            bv = ws.cell(row=rr, column=COL_LABEL).value
            if is_continuation(bv):
                if current_grp is not None:
                    current_grp["rows"].append(rr)
            else:
                current_grp = {"name": str(bv).strip(), "rows": [rr]}
                groups.append(current_grp)

        if not groups:
            r = block_end + 1
            continue

        ctrl_rows   = groups[0]["rows"]
        ctrl_label  = groups[0]["name"]
        drug_rows   = []
        drug_labels = []
        for g in groups[1:]:
            drug_rows.extend(g["rows"])
            drug_labels.append(g["name"])

        blocks.append({
            "organ":          organ,
            "hk_name":        hk_name   or "HK Gene",
            "test_name":      test_name or "Test Gene",
            "first_data_row": block_start,
            "last_data_row":  block_end,
            "ctrl_rows":      ctrl_rows,
            "ctrl_label":     ctrl_label,
            "drug_rows":      drug_rows,
            "drug_labels":    drug_labels,
        })

        r = block_end + 1

    return blocks


# ── Numerical computation ─────────────────────────────────────────────────────

def compute_values(ws, blocks) -> list:
    """
    Compute delta and relative expression using raw float values.
    Augments each block dict in-place and returns the list.
    """
    for blk in blocks:
        all_rows = blk["ctrl_rows"] + blk["drug_rows"]

        deltas = {}
        for row in all_rows:
            hk   = float(ws.cell(row=row, column=COL_HK).value)
            test = float(ws.cell(row=row, column=COL_TEST).value)
            deltas[row] = test - hk

        ctrl_deltas   = [deltas[r] for r in blk["ctrl_rows"]]
        avg_ctrl      = sum(ctrl_deltas) / len(ctrl_deltas)
        rel           = {r: 2 ** (avg_ctrl - deltas[r]) for r in all_rows}

        ctrl_rel = [rel[r] for r in blk["ctrl_rows"]]
        drug_rel = [rel[r] for r in blk["drug_rows"]]

        blk["deltas"]         = deltas
        blk["avg_ctrl_delta"] = avg_ctrl
        blk["ctrl_rel"]       = ctrl_rel
        blk["drug_rel"]       = drug_rel
        blk["ctrl_mean_re"]   = float(np.mean(ctrl_rel)) if ctrl_rel else 0.0
        blk["drug_mean_re"]   = float(np.mean(drug_rel)) if drug_rel else 0.0

    return blocks


# ── Output workbook builder ───────────────────────────────────────────────────

def build_output_workbook(ws_source, blocks) -> bytes:
    """
    Clone the source worksheet, insert a legend row at the top, reuse or
    insert a header row per organ block, then write formulas into cols E & F.

    Key design rules that avoid the off-by-one bugs:
      1. Insert the legend row FIRST so every subsequent row number already
         includes that +1 offset before any formula strings are written.
      2. Check whether a text header row already exists immediately above
         each block's first data row. If it does, reuse it; don't insert a
         duplicate.
      3. When inserting new header rows, process blocks bottom-to-top so
         inserting a row for a lower block doesn't shift the row indices of
         upper blocks that we've already accounted for.
      4. Compute ALL final row numbers BEFORE writing any formula strings,
         then write every formula in a single pass using those final numbers.
    """

    # ── Step 1: clone the source workbook ────────────────────────────────────
    src_buf = io.BytesIO()
    ws_source.parent.save(src_buf)
    wb_out = load_workbook(io.BytesIO(src_buf.getvalue()))
    ws_out = wb_out.active

    # ── Step 2: insert legend row at row 1 (everything else shifts +1) ───────
    ws_out.insert_rows(1)
    ws_out["A1"] = "Legend:"
    ws_out["B1"] = "Gray = Control   |   Red = Drug-treated"
    ws_out["A1"].font = Font(bold=True, name="Arial", size=10)
    ws_out["B1"].font = Font(italic=True, name="Arial", size=10)

    # All original row numbers now have +1 applied.
    # Adjust block row references to match the post-legend state.
    for blk in blocks:
        blk["_first"]     = blk["first_data_row"] + 1
        blk["_last"]      = blk["last_data_row"]  + 1
        blk["_ctrl_cur"]  = [r + 1 for r in blk["ctrl_rows"]]
        blk["_drug_cur"]  = [r + 1 for r in blk["drug_rows"]]

    # ── Step 3: decide whether each block needs a new header row ─────────────
    # A block already has a header if the row immediately above its first data
    # row is a non-numeric text row (i.e. the Cyclo / CD19-3 label row).
    for blk in blocks:
        row_above = blk["_first"] - 1
        if row_above >= 1:
            c_above = ws_out.cell(row=row_above, column=COL_HK).value
            d_above = ws_out.cell(row=row_above, column=COL_TEST).value
            has_existing = (
                isinstance(c_above, str) and c_above.strip() != "" and
                not is_numeric(c_above)
            ) or (
                isinstance(d_above, str) and d_above.strip() != "" and
                not is_numeric(d_above)
            )
        else:
            has_existing = False

        blk["_has_existing_header"] = has_existing
        blk["_needs_insert"]        = not has_existing

    # ── Step 4: insert missing header rows, bottom-to-top ────────────────────
    # Processing bottom-to-top means inserting into block N doesn't shift
    # the "_first" row numbers of blocks 0..N-1 which we haven't touched yet.
    sorted_desc = sorted(blocks, key=lambda b: b["_first"], reverse=True)
    sorted_asc  = sorted(blocks, key=lambda b: b["_first"])

    # Track how many rows each block needs to shift due to insertions BELOW it.
    # Since we go bottom-to-top, when we insert for block N none of the
    # blocks above it have been processed yet — their "_first" etc. are still
    # correct. We just need to update blocks that sit ABOVE the insertion point
    # after each insertion. Easiest: accumulate offset and apply at the end.

    # Collect insertion points first (while row numbers are still pristine)
    insertions = []   # list of row numbers where we will insert
    for blk in sorted_desc:
        if blk["_needs_insert"]:
            insertions.append(blk["_first"])   # insert immediately before first data row

    # Now actually insert, bottom-to-top (sorted_desc already is bottom-to-top)
    cumulative = 0   # counts insertions done so far (above already-processed blocks)
    # We process bottom-to-top, so each insertion doesn't affect the blocks
    # above (which we haven't processed yet).  But it DOES shift the blocks
    # below that we already processed — those are done, so no problem.

    # Reset cumulative; go bottom-to-top and track shifts for blocks above.
    # Simpler approach: do all inserts bottom-to-top and update every block
    # above the insertion point each time.

    for blk in sorted_desc:
        if not blk["_needs_insert"]:
            # Still need to record where its header is (the existing row above)
            blk["_hdr_final"]  = blk["_first"] - 1
            blk["_ctrl_final"] = blk["_ctrl_cur"][:]
            blk["_drug_final"] = blk["_drug_cur"][:]
            continue

        insert_at = blk["_first"]   # insert a new row HERE; data shifts to insert_at+1

        ws_out.insert_rows(insert_at)

        # Write header labels into the new row
        ws_out.cell(row=insert_at, column=COL_HK).value    = blk["hk_name"]
        ws_out.cell(row=insert_at, column=COL_TEST).value  = blk["test_name"]
        ws_out.cell(row=insert_at, column=COL_DELTA).value = "delta"
        ws_out.cell(row=insert_at, column=COL_REL).value   = "Avg Ctrl Δ / Rel Expr"

        # This block's header and data rows after the insertion
        blk["_hdr_final"]  = insert_at
        blk["_ctrl_final"] = [r + 1 for r in blk["_ctrl_cur"]]
        blk["_drug_final"] = [r + 1 for r in blk["_drug_cur"]]

        # Update all blocks that sit ABOVE this insertion point (i.e. smaller
        # _first value) — they haven't been processed yet in this loop so their
        # _ctrl_cur / _drug_cur / _first still need shifting.
        for other in blocks:
            if other is blk:
                continue
            if other["_first"] < insert_at:
                # Block sits above the insertion — not affected
                pass
            # Block sits below or at insertion — already processed (done), skip
            # Block sits at same point — shouldn't happen

        # Also update blocks that are ABOVE (first < insert_at) in the sorted_asc
        # list that we haven't processed yet (they are earlier in sorted_desc
        # so they come later in sorted_desc iteration — haven't been touched).
        # Because we go bottom-to-top, blocks above have SMALLER _first values,
        # so they are NOT shifted by an insertion at insert_at. Correct: inserting
        # at row 5 doesn't change rows 1-4. ✓

    # For blocks that had existing headers (no insert), set their finals now
    # (after all insertions are done, so we know the true final positions).
    # We need to figure out how many insertions happened ABOVE each such block.
    # Recalculate by scanning sorted_asc and tracking cumulative inserted rows.

    cumulative_above = {}
    running = 0
    # sorted_asc is in ascending order of original _first
    # We inserted for blocks in sorted_desc order.
    # Count how many insertions happened at rows BELOW each block.
    all_insert_rows = [b["_first"] for b in blocks if b["_needs_insert"]]

    for blk in sorted_asc:
        # How many insertions are at rows < blk["_first"]?
        # (inserting at a row >= blk["_first"] shifts blk upward)
        shifts = sum(1 for ins_row in all_insert_rows if ins_row <= blk["_first"])
        # But for blocks that needed insert, _hdr_final / _ctrl_final are already set
        if not blk["_needs_insert"]:
            # existing header row was at _first - 1; now shifted by `shifts`
            blk["_hdr_final"]  = blk["_first"] - 1 + shifts
            blk["_ctrl_final"] = [r + shifts for r in blk["_ctrl_cur"]]
            blk["_drug_final"] = [r + shifts for r in blk["_drug_cur"]]

    # ── Step 5: write all formulas using the final, correct row numbers ───────
    for blk in sorted_asc:
        h       = blk["_hdr_final"]
        all_out = blk["_ctrl_final"] + blk["_drug_final"]

        # Delta: E{row} = D{row} - C{row}
        for row in all_out:
            ws_out.cell(row=row, column=COL_DELTA).value = (
                f"={cl(COL_TEST)}{row}-{cl(COL_HK)}{row}"
            )

        # Average control delta in header row col F
        ctrl_refs = ",".join(f"{cl(COL_DELTA)}{r}" for r in blk["_ctrl_final"])
        ws_out.cell(row=h, column=COL_REL).value = f"=AVERAGE({ctrl_refs})"

        # Relative expression: F{row} = 2^(F{h} - E{row})
        for row in all_out:
            ws_out.cell(row=row, column=COL_REL).value = (
                f"=2^({cl(COL_REL)}{h}-{cl(COL_DELTA)}{row})"
            )

    # ── Step 6: style ─────────────────────────────────────────────────────────
    _style_output(ws_out, sorted_asc)

    out_buf = io.BytesIO()
    wb_out.save(out_buf)
    return out_buf.getvalue()


# ── Styling ───────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

_CENTER = Alignment(horizontal="center", vertical="center")

def _style_output(ws, blocks):
    """Style header and data rows. Uses _hdr_final / _ctrl_final / _drug_final
    which already incorporate the legend row offset and any inserted header rows."""
    ws.column_dimensions[cl(COL_ORGAN)].width = 10
    ws.column_dimensions[cl(COL_LABEL)].width = 22
    ws.column_dimensions[cl(COL_HK)].width    = 14
    ws.column_dimensions[cl(COL_TEST)].width   = 14
    ws.column_dimensions[cl(COL_DELTA)].width  = 12
    ws.column_dimensions[cl(COL_REL)].width    = 22

    hdr_fill  = _fill(HDR_HEX)
    ctrl_fill = _fill(CTRL_HEX)
    drug_fill = _fill(DRUG_HEX)
    border    = _border()
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    std_font  = Font(name="Arial", size=10)

    for blk in blocks:
        h         = blk["_hdr_final"]
        ctrl_rows = blk["_ctrl_final"]
        drug_rows = blk["_drug_final"]

        for col in range(1, COL_REL + 1):
            cell = ws.cell(row=h, column=col)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = _CENTER; cell.border = border

        for row in ctrl_rows:
            for col in range(1, COL_REL + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = std_font; cell.fill = ctrl_fill
                cell.alignment = _CENTER; cell.border = border

        for row in drug_rows:
            for col in range(1, COL_REL + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = std_font; cell.fill = drug_fill
                cell.alignment = _CENTER; cell.border = border


# ── Chart renderer ────────────────────────────────────────────────────────────

def render_chart(blocks: list) -> bytes:
    organs     = [b["organ"]        for b in blocks]
    ctrl_means = [b["ctrl_mean_re"] for b in blocks]
    drug_means = [b["drug_mean_re"] for b in blocks]
    n          = len(organs)

    ctrl_label = blocks[0]["ctrl_label"]             if blocks else "Control"
    drug_label = ", ".join(blocks[0]["drug_labels"]) if blocks else "Treatment"

    group_spacing = 1.4
    bar_width     = 0.22
    half_span     = bar_width / 2 + 0.05
    x             = np.arange(n) * group_spacing

    fig, ax = plt.subplots(figsize=(max(5, 2.4 * n + 1.8), 5.5))
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    ax.bar(x - half_span, ctrl_means, width=bar_width, color=CTRL_MPL, zorder=3)
    ax.bar(x + half_span, drug_means, width=bar_width, color=DRUG_MPL, zorder=3)

    ax.set_xticks(x)
    ax.set_xticklabels(organs, fontsize=13)
    ax.set_ylabel("Relative Expression", fontsize=12)
    ax.set_ylim(bottom=0)
    ax.yaxis.grid(True, color="#E8E8E8", linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#AAAAAA")
    ax.spines["bottom"].set_color("#AAAAAA")
    ax.tick_params(axis="both", which="both", length=0, labelsize=11)
    ax.set_xlim(-group_spacing * 0.55,
                (n - 1) * group_spacing + group_spacing * 0.55)

    ctrl_patch = mpatches.Patch(color=CTRL_MPL, label=ctrl_label)
    drug_patch = mpatches.Patch(color=DRUG_MPL, label=drug_label)
    ax.legend(handles=[ctrl_patch, drug_patch], loc="upper center",
              bbox_to_anchor=(0.5, 1.13), ncol=2, frameon=False,
              fontsize=11, handlelength=1.0, handleheight=1.0)

    plt.tight_layout(rect=[0, 0, 1, 0.95])
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=180, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close()
    buf.seek(0)
    return buf.read()


# ── Top-level pipeline ────────────────────────────────────────────────────────

def run_pipeline(file_bytes: bytes) -> dict:
    result = {
        "xlsx_bytes": None, "chart_bytes": None,
        "chart_data": [],   "blocks": [],   "errors": "",
    }

    try:
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
    except Exception as e:
        result["errors"] = f"Could not open file: {e}"
        return result

    blocks = parse_blocks(ws)
    if not blocks:
        result["errors"] = (
            "No data blocks detected. Make sure columns C and D contain "
            "numeric values and each organ group is separated by a blank row."
        )
        return result

    try:
        blocks = compute_values(ws, blocks)
    except Exception as e:
        result["errors"] = f"Calculation error: {e}"
        return result

    result["blocks"] = blocks
    result["chart_data"] = [
        {
            "organ":       b["organ"],
            "ctrl_mean":   b["ctrl_mean_re"],
            "drug_mean":   b["drug_mean_re"],
            "ctrl_label":  b["ctrl_label"],
            "drug_labels": b["drug_labels"],
        }
        for b in blocks
    ]

    try:
        result["xlsx_bytes"] = build_output_workbook(ws, blocks)
    except Exception as e:
        result["errors"] = f"Excel output error: {e}"
        return result

    try:
        result["chart_bytes"] = render_chart(blocks)
    except Exception as e:
        result["errors"] = f"Chart rendering failed: {e}"
        return result

    return result

"""
process_gene_expression.py
---------------------------
Core qPCR analysis logic — pure Python, no LibreOffice dependency.

DATA MODEL
----------
Each organ block produces one dict with this shape:

  {
    "organ"         : str,
    "hk_name"       : str,
    "test_name"     : str,
    "first_data_row": int,
    "last_data_row" : int,

    # The control group (always the first named group in the block)
    "ctrl_label"    : str,
    "ctrl_rows"     : [int],          # 1-based source row indices

    # Every drug group, kept separate, in the order they appear
    "drug_groups"   : [
        {
          "label": str,
          "rows" : [int],             # 1-based source row indices
          "rel"  : [float],           # computed relative expression values
          "mean_re": float,           # mean across replicates
        },
        ...
    ],

    # Control-side computed values (added by compute_values)
    "deltas"        : {row: float},
    "avg_ctrl_delta": float,
    "ctrl_rel"      : [float],
    "ctrl_mean_re"  : float,
  }

Spreadsheet format (per organ block):
  Optional text header row above the data rows:
    col A = organ name, col C = HK gene name, col D = test gene name
  Data rows:
    col A = organ name (first row only, or in text header row)
    col B = sample label  (blank / "cont..." = replicate of current group)
    col C = housekeeping gene value  (float)
    col D = test gene value          (float)
  Blocks are separated by rows where C and D are both non-numeric.

Grouping rules:
  • The FIRST named group (non-"cont" col B) in a block = control.
  • Any row whose col B is blank or starts with "cont" (case-insensitive)
    is a replicate of the currently open group.
  • Every subsequent named group = its own independent drug group.
"""

import io
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Column indices (1-based) ──────────────────────────────────────────────────
COL_ORGAN = 1   # A
COL_LABEL = 2   # B
COL_HK    = 3   # C  housekeeping gene
COL_TEST  = 4   # D  test gene
COL_DELTA = 5   # E  computed: test − HK
COL_REL   = 6   # F  computed: 2^(avg_ctrl_delta − delta)

# ── Style constants ───────────────────────────────────────────────────────────
CTRL_HEX = "D9D9D9"   # light gray fill for control rows
DRUG_HEX = "F4A99A"   # light red fill for drug rows
HDR_HEX  = "2F4F8F"   # dark blue fill for header rows
CTRL_MPL = "#C8C8C8"
DRUG_MPL = "#E04030"


# ── Utilities ─────────────────────────────────────────────────────────────────

def cl(n: int) -> str:
    return get_column_letter(n)

def is_numeric(v) -> bool:
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False

def is_continuation(label) -> bool:
    """True when a col-B value marks a replicate of the current group."""
    if label is None:
        return True
    s = str(label).strip().lower()
    return s == "" or s.startswith("cont")


# ── Block parser ──────────────────────────────────────────────────────────────

def parse_blocks(ws) -> list:
    """
    Scan the worksheet and return one dict per organ block.
    See module docstring for the full dict shape.
    """
    max_row      = ws.max_row
    numeric_rows = set()

    for r in range(1, max_row + 1):
        if (is_numeric(ws.cell(row=r, column=COL_HK).value) and
                is_numeric(ws.cell(row=r, column=COL_TEST).value)):
            numeric_rows.add(r)

    blocks  = []
    visited = set()
    r       = 1

    while r <= max_row:
        if r not in numeric_rows or r in visited:
            r += 1
            continue

        # ── find contiguous block of numeric rows ──────────────────────────
        block_start = r
        block_end   = r
        while block_end + 1 in numeric_rows:
            block_end += 1
        for rr in range(block_start, block_end + 1):
            visited.add(rr)

        # ── look for organ name / gene names in rows just above the data ──
        organ     = None
        hk_name   = None
        test_name = None

        for look in range(block_start - 1, max(0, block_start - 4), -1):
            a = ws.cell(row=look, column=COL_ORGAN).value
            c = ws.cell(row=look, column=COL_HK).value
            d = ws.cell(row=look, column=COL_TEST).value
            if a is not None and str(a).strip():
                organ = str(a).strip()
            if isinstance(c, str) and c.strip() and not is_numeric(c):
                hk_name = c.strip()
            if isinstance(d, str) and d.strip() and not is_numeric(d):
                test_name = d.strip()
            if organ:
                break

        # Fallback: organ name may live inside the data rows (col A)
        if organ is None:
            for rr in range(block_start, block_end + 1):
                a = ws.cell(row=rr, column=COL_ORGAN).value
                if a is not None and str(a).strip():
                    organ = str(a).strip()
                    break

        if organ is None:
            organ = f"Block_{block_start}"

        # ── group data rows by sample label ───────────────────────────────
        groups      = []   # [{"name": str, "rows": [int]}]
        current_grp = None

        for rr in range(block_start, block_end + 1):
            bv = ws.cell(row=rr, column=COL_LABEL).value
            if is_continuation(bv):
                if current_grp is not None:
                    current_grp["rows"].append(rr)
                # Stray continuation before any named group → ignore
            else:
                current_grp = {"name": str(bv).strip(), "rows": [rr]}
                groups.append(current_grp)

        if not groups:
            r = block_end + 1
            continue

        # First group = control; all remaining = independent drug groups
        ctrl_group  = groups[0]
        drug_groups = [
            {"label": g["name"], "rows": g["rows"], "rel": [], "mean_re": 0.0}
            for g in groups[1:]
        ]

        blocks.append({
            "organ":          organ,
            "hk_name":        hk_name   or "HK Gene",
            "test_name":      test_name or "Test Gene",
            "first_data_row": block_start,
            "last_data_row":  block_end,
            "ctrl_label":     ctrl_group["name"],
            "ctrl_rows":      ctrl_group["rows"],
            "drug_groups":    drug_groups,
            # Computed fields added later by compute_values():
            "deltas":         {},
            "avg_ctrl_delta": 0.0,
            "ctrl_rel":       [],
            "ctrl_mean_re":   0.0,
        })

        r = block_end + 1

    return blocks


# ── Numerical computation ─────────────────────────────────────────────────────

def compute_values(ws, blocks: list) -> list:
    """
    Compute delta and relative expression for every row in every block.
    Updates each block (and each drug_group within it) in-place.
    Returns the mutated blocks list.
    """
    for blk in blocks:
        all_drug_rows = [r for dg in blk["drug_groups"] for r in dg["rows"]]
        all_rows      = blk["ctrl_rows"] + all_drug_rows

        # delta = test_gene − housekeeping_gene  (for every replicate row)
        deltas = {}
        for row in all_rows:
            hk   = float(ws.cell(row=row, column=COL_HK).value)
            test = float(ws.cell(row=row, column=COL_TEST).value)
            deltas[row] = test - hk

        # average control delta (goes into the header row col F in the output)
        ctrl_deltas   = [deltas[r] for r in blk["ctrl_rows"]]
        avg_ctrl      = sum(ctrl_deltas) / len(ctrl_deltas)

        # relative expression = 2^(avg_ctrl_delta − delta)
        rel_expr = {r: 2.0 ** (avg_ctrl - deltas[r]) for r in all_rows}

        # store control-side results
        ctrl_rel = [rel_expr[r] for r in blk["ctrl_rows"]]
        blk["deltas"]         = deltas
        blk["avg_ctrl_delta"] = avg_ctrl
        blk["ctrl_rel"]       = ctrl_rel
        blk["ctrl_mean_re"]   = float(np.mean(ctrl_rel)) if ctrl_rel else 0.0

        # store per-drug-group results
        for dg in blk["drug_groups"]:
            dg_rel        = [rel_expr[r] for r in dg["rows"]]
            dg["rel"]     = dg_rel
            dg["mean_re"] = float(np.mean(dg_rel)) if dg_rel else 0.0

    return blocks


# ── Output workbook builder ───────────────────────────────────────────────────

def build_output_workbook(ws_source, blocks: list) -> bytes:
    """
    Clone the source worksheet, insert a legend row at the top, reuse or
    insert a header row per organ block, then write delta and relative-
    expression formulas into cols E and F.

    Insertion order:
      1. Legend row at row 1 (all original row numbers shift +1 immediately).
      2. Header rows for blocks that don't have one, processed bottom-to-top
         so earlier insertions don't invalidate later row references.
      3. Formula strings written only AFTER all insertions, using the final
         post-insertion row numbers, so no off-by-one errors occur.
    """

    # ── clone ─────────────────────────────────────────────────────────────────
    src_buf = io.BytesIO()
    ws_source.parent.save(src_buf)
    wb_out = load_workbook(io.BytesIO(src_buf.getvalue()))
    ws_out = wb_out.active

    # ── Step 1: legend row ────────────────────────────────────────────────────
    ws_out.insert_rows(1)
    ws_out["A1"] = "Legend:"
    ws_out["B1"] = "Gray = Control   |   Red = Drug-treated"
    ws_out["A1"].font = Font(bold=True, name="Arial", size=10)
    ws_out["B1"].font = Font(italic=True, name="Arial", size=10)

    # Shift all block row references by +1 for the legend row.
    for blk in blocks:
        blk["_first"]    = blk["first_data_row"] + 1
        blk["_last"]     = blk["last_data_row"]  + 1
        blk["_ctrl_cur"] = [r + 1 for r in blk["ctrl_rows"]]
        blk["_dgs_cur"]  = [
            {"label": dg["label"], "rows": [r + 1 for r in dg["rows"]]}
            for dg in blk["drug_groups"]
        ]

    # ── Step 2: detect which blocks need a new header row inserted ────────────
    for blk in blocks:
        row_above = blk["_first"] - 1
        if row_above >= 1:
            c_val = ws_out.cell(row=row_above, column=COL_HK).value
            d_val = ws_out.cell(row=row_above, column=COL_TEST).value
            existing = (
                (isinstance(c_val, str) and c_val.strip() and not is_numeric(c_val)) or
                (isinstance(d_val, str) and d_val.strip() and not is_numeric(d_val))
            )
        else:
            existing = False
        blk["_has_hdr"]     = existing
        blk["_needs_insert"] = not existing

    # collect which original _first rows will have a header inserted
    all_insert_rows = [b["_first"] for b in blocks if b["_needs_insert"]]

    # ── Step 3: insert missing header rows, bottom-to-top ─────────────────────
    sorted_desc = sorted(blocks, key=lambda b: b["_first"], reverse=True)
    sorted_asc  = sorted(blocks, key=lambda b: b["_first"])

    for blk in sorted_desc:
        if not blk["_needs_insert"]:
            continue
        insert_at = blk["_first"]
        ws_out.insert_rows(insert_at)
        ws_out.cell(row=insert_at, column=COL_HK).value    = blk["hk_name"]
        ws_out.cell(row=insert_at, column=COL_TEST).value  = blk["test_name"]
        ws_out.cell(row=insert_at, column=COL_DELTA).value = "delta"
        ws_out.cell(row=insert_at, column=COL_REL).value   = "Avg Ctrl Δ / Rel Expr"
        blk["_hdr_inserted_at"] = insert_at

    # ── Step 4: compute final row numbers for every block ─────────────────────
    # For each block, count how many header-row insertions occurred at rows
    # <= that block's _first (those shift its data rows upward).
    for blk in sorted_asc:
        shifts = sum(1 for ins in all_insert_rows if ins <= blk["_first"])

        if blk["_needs_insert"]:
            # Its own insertion is included in shifts; header is at _first + shifts - 1
            # (the insert pushed everything at _first upward by 1, then data starts 1 below)
            blk["_hdr_final"]  = blk["_first"] + shifts - 1
            blk["_ctrl_final"] = [r + shifts for r in blk["_ctrl_cur"]]
            blk["_dgs_final"]  = [
                {"label": dg["label"], "rows": [r + shifts for r in dg["rows"]]}
                for dg in blk["_dgs_cur"]
            ]
        else:
            # No insertion for this block; existing header at _first - 1, shifted by shifts
            blk["_hdr_final"]  = blk["_first"] - 1 + shifts
            blk["_ctrl_final"] = [r + shifts for r in blk["_ctrl_cur"]]
            blk["_dgs_final"]  = [
                {"label": dg["label"], "rows": [r + shifts for r in dg["rows"]]}
                for dg in blk["_dgs_cur"]
            ]

    # ── Step 5: write formulas using final row numbers ─────────────────────────
    for blk in sorted_asc:
        h        = blk["_hdr_final"]
        all_rows = (
            blk["_ctrl_final"] +
            [r for dg in blk["_dgs_final"] for r in dg["rows"]]
        )

        # E: delta = test − HK
        for row in all_rows:
            ws_out.cell(row=row, column=COL_DELTA).value = (
                f"={cl(COL_TEST)}{row}-{cl(COL_HK)}{row}"
            )

        # F header: average control delta
        ctrl_refs = ",".join(f"{cl(COL_DELTA)}{r}" for r in blk["_ctrl_final"])
        ws_out.cell(row=h, column=COL_REL).value = f"=AVERAGE({ctrl_refs})"

        # F data rows: 2^(avg_ctrl_delta − delta)
        for row in all_rows:
            ws_out.cell(row=row, column=COL_REL).value = (
                f"=2^({cl(COL_REL)}{h}-{cl(COL_DELTA)}{row})"
            )

    # ── Step 6: style ──────────────────────────────────────────────────────────
    _style_output(ws_out, sorted_asc)

    out_buf = io.BytesIO()
    wb_out.save(out_buf)
    return out_buf.getvalue()


# ── Styling ───────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _border() -> Border:
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

_CENTER = Alignment(horizontal="center", vertical="center")


def _style_output(ws, blocks: list) -> None:
    ws.column_dimensions[cl(COL_ORGAN)].width = 10
    ws.column_dimensions[cl(COL_LABEL)].width = 22
    ws.column_dimensions[cl(COL_HK)].width    = 14
    ws.column_dimensions[cl(COL_TEST)].width  = 14
    ws.column_dimensions[cl(COL_DELTA)].width = 12
    ws.column_dimensions[cl(COL_REL)].width   = 22

    hdr_fill  = _fill(HDR_HEX)
    ctrl_fill = _fill(CTRL_HEX)
    drug_fill = _fill(DRUG_HEX)
    border    = _border()
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    std_font  = Font(name="Arial", size=10)

    for blk in blocks:
        h = blk["_hdr_final"]

        # Header row
        for col in range(1, COL_REL + 1):
            cell = ws.cell(row=h, column=col)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = _CENTER; cell.border = border

        # Control rows
        for row in blk["_ctrl_final"]:
            for col in range(1, COL_REL + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = std_font; cell.fill = ctrl_fill
                cell.alignment = _CENTER; cell.border = border

        # Drug rows (all drug groups, same red fill)
        for dg in blk["_dgs_final"]:
            for row in dg["rows"]:
                for col in range(1, COL_REL + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = std_font; cell.fill = drug_fill
                    cell.alignment = _CENTER; cell.border = border


# ── Chart renderer ────────────────────────────────────────────────────────────

def render_chart(blocks: list) -> bytes:
    """
    Render a grouped bar chart with one bar per group (control + each drug group)
    per organ. All drug bars are the same red; control bars are gray.
    The legend lists every unique group label found across all blocks.

    Bar layout within each organ cluster:
      [ctrl] [drug_0] [drug_1] ...
    Spacing is kept consistent even when different organs have different numbers
    of drug groups — organs with fewer groups simply have shorter clusters.
    """

    # ── collect all unique group labels in order of first appearance ──────────
    ctrl_labels   = []
    drug_labels   = []   # ordered list of every distinct drug-group name seen

    for blk in blocks:
        if blk["ctrl_label"] not in ctrl_labels:
            ctrl_labels.append(blk["ctrl_label"])
        for dg in blk["drug_groups"]:
            if dg["label"] not in drug_labels:
                drug_labels.append(dg["label"])

    # Use the first (and usually only) control label for the legend
    ctrl_legend_label = ctrl_labels[0] if ctrl_labels else "Control"

    # Total number of possible bar slots per organ group:
    # 1 (control) + number of distinct drug groups across all blocks
    n_drug_slots  = len(drug_labels)
    n_slots       = 1 + n_drug_slots   # bars per organ cluster

    organs = [blk["organ"] for blk in blocks]
    n_organs = len(organs)

    # ── bar geometry ──────────────────────────────────────────────────────────
    bar_width     = 0.18
    bar_gap       = 0.06   # gap between adjacent bars within a cluster
    cluster_width = n_slots * bar_width + (n_slots - 1) * bar_gap
    group_spacing = cluster_width + 0.55   # centre-to-centre distance between organs

    x = np.arange(n_organs) * group_spacing   # centre of each organ cluster

    # Offset of each slot within a cluster, centred on x
    # Slot 0 = control, slots 1..n_drug_slots = drug groups in order
    slot_offsets = np.array([
        -cluster_width / 2 + i * (bar_width + bar_gap) + bar_width / 2
        for i in range(n_slots)
    ])

    # ── figure ────────────────────────────────────────────────────────────────
    fig_w = max(5.5, group_spacing * n_organs + 1.5)
    fig_h = 5.5 + 0.25 * max(0, n_drug_slots - 1)   # taller legend if many groups
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    # ── draw bars ─────────────────────────────────────────────────────────────
    # Build a lookup: drug_label → slot index (1-based)
    drug_slot = {label: i + 1 for i, label in enumerate(drug_labels)}

    for i, blk in enumerate(blocks):
        # Control bar (slot 0)
        ax.bar(
            x[i] + slot_offsets[0],
            blk["ctrl_mean_re"],
            width=bar_width,
            color=CTRL_MPL,
            zorder=3,
        )

        # Drug bars — only the slots this organ actually has
        for dg in blk["drug_groups"]:
            slot = drug_slot[dg["label"]]
            ax.bar(
                x[i] + slot_offsets[slot],
                dg["mean_re"],
                width=bar_width,
                color=DRUG_MPL,
                zorder=3,
            )

    # ── legend ────────────────────────────────────────────────────────────────
    legend_handles = [mpatches.Patch(color=CTRL_MPL, label=ctrl_legend_label)]
    for label in drug_labels:
        legend_handles.append(mpatches.Patch(color=DRUG_MPL, label=label))

    n_legend_cols = min(len(legend_handles), 4)   # max 4 per row
    ax.legend(
        handles=legend_handles,
        loc="upper center",
        bbox_to_anchor=(0.5, 1.13),
        ncol=n_legend_cols,
        frameon=False,
        fontsize=11,
        handlelength=1.0,
        handleheight=1.0,
    )

    # ── axes formatting ───────────────────────────────────────────────────────
    ax.set_xticks(x)
    ax.set_xticklabels(organs, fontsize=13)
    ax.set_ylabel("Relative Expression", fontsize=12)
    ax.set_ylim(bottom=0)
    ax.yaxis.grid(True, color="#E8E8E8", linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#AAAAAA")
    ax.spines["bottom"].set_color("#AAAAAA")
    ax.tick_params(axis="both", which="both", length=0, labelsize=11)
    margin = group_spacing * 0.45
    ax.set_xlim(x[0] - cluster_width / 2 - margin,
                x[-1] + cluster_width / 2 + margin)

    plt.tight_layout(rect=[0, 0, 1, 0.93])
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=180, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close()
    buf.seek(0)
    return buf.read()


# ── Top-level pipeline ────────────────────────────────────────────────────────

def run_pipeline(file_bytes: bytes) -> dict:
    """
    Accept raw .xlsx bytes, run the full analysis, return:
    {
        "xlsx_bytes"  : bytes
        "chart_bytes" : bytes
        "chart_data"  : list of per-organ dicts (see below)
        "blocks"      : list of fully-computed block dicts
        "errors"      : str  (non-empty if something went wrong)
    }

    Each entry in chart_data:
    {
        "organ"      : str,
        "ctrl_label" : str,
        "ctrl_mean"  : float,
        "drug_groups": [{"label": str, "mean_re": float}, ...]
    }
    """
    result = {
        "xlsx_bytes":  None,
        "chart_bytes": None,
        "chart_data":  [],
        "blocks":      [],
        "errors":      "",
    }

    # Load
    try:
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
    except Exception as e:
        result["errors"] = f"Could not open file: {e}"
        return result

    # Parse
    blocks = parse_blocks(ws)
    if not blocks:
        result["errors"] = (
            "No data blocks detected. Make sure columns C and D contain "
            "numeric values and each organ group is separated by a blank row."
        )
        return result

    # Compute
    try:
        blocks = compute_values(ws, blocks)
    except Exception as e:
        result["errors"] = f"Calculation error: {e}"
        return result

    result["blocks"] = blocks

    # Build chart_data (one entry per organ, drug_groups kept separate)
    result["chart_data"] = [
        {
            "organ":       blk["organ"],
            "ctrl_label":  blk["ctrl_label"],
            "ctrl_mean":   blk["ctrl_mean_re"],
            "drug_groups": [
                {"label": dg["label"], "mean_re": dg["mean_re"]}
                for dg in blk["drug_groups"]
            ],
        }
        for blk in blocks
    ]

    # Build Excel output
    try:
        result["xlsx_bytes"] = build_output_workbook(ws, blocks)
    except Exception as e:
        result["errors"] = f"Excel output error: {e}"
        return result

    # Render chart
    try:
        result["chart_bytes"] = render_chart(blocks)
    except Exception as e:
        result["errors"] = f"Chart rendering failed: {e}"
        return result

    return result

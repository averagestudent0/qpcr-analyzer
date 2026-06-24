"""
process_gene_expression.py
---------------------------
Core qPCR analysis logic. All functions are pure / side-effect-free:
they accept workbook / worksheet objects and return results or mutate
the worksheet in place. No file I/O happens here — the Streamlit app
handles reading and writing.

Spreadsheet format expected (per organ block):
  Header row  : [blank] [blank] <HK_gene>  <test_gene>  (cols A–D, row label in E is "delta" if pre-existing)
  Data rows   : [organ] <label> <HK_val>   <test_val>
  ...more replicates (col A blank, col B can be blank or "cont…")
  [blank row] → end of block, next block starts

Control vs drug grouping:
  The FIRST named group per organ block is the control.
  A "cont" (case-insensitive) or blank col-B row is a replicate of the
  current group — it does NOT start a new group.
  Every subsequent named group is a drug / treatment group.
"""

import io
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Column indices (1-based) ──────────────────────────────────────────────────
COL_ORGAN   = 1   # A
COL_LABEL   = 2   # B
COL_HK      = 3   # C  housekeeping gene
COL_TEST    = 4   # D  gene of interest
COL_DELTA   = 5   # E  written by this script
COL_REL     = 6   # F  written by this script

# ── Style constants ───────────────────────────────────────────────────────────
CTRL_HEX   = "D9D9D9"
DRUG_HEX   = "F4A99A"
HDR_HEX    = "2F4F8F"
CTRL_MPL   = "#C8C8C8"
DRUG_MPL   = "#E04030"


# ── Helpers ───────────────────────────────────────────────────────────────────

def cl(col_index: int) -> str:
    return get_column_letter(col_index)


def is_continuation(label) -> bool:
    """True if col-B value means 'another replicate of the current group'."""
    if label is None:
        return True
    s = str(label).strip().lower()
    return s == "" or s.startswith("cont")


# ── Block parser ──────────────────────────────────────────────────────────────

def parse_blocks(ws) -> list[dict]:
    """
    Scan the worksheet and return one dict per organ block:
    {
        organ       : str
        header_row  : int   (1-based row that will hold the delta/avg header)
        ctrl_rows   : [int]
        ctrl_label  : str
        drug_rows   : [int]
        drug_labels : [str]
        hk_name     : str   (col C header if present, else None)
        test_name   : str   (col D header if present, else None)
    }

    A "header row" is the row immediately BEFORE the first data row of a block.
    If the original sheet already has a header row (col E == "delta"), we reuse it.
    Otherwise we insert one.
    """
    blocks = []
    max_row = ws.max_row

    # Identify existing header rows (col E == "delta")
    header_row_set: set[int] = set()
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=COL_DELTA).value
        if isinstance(v, str) and v.strip().lower() == "delta":
            header_row_set.add(r)

    # Walk rows looking for data blocks
    r = 1
    while r <= max_row:
        a = ws.cell(row=r, column=COL_ORGAN).value
        b = ws.cell(row=r, column=COL_LABEL).value
        c = ws.cell(row=r, column=COL_HK).value
        d = ws.cell(row=r, column=COL_TEST).value

        # Skip rows that are header rows or fully blank
        if r in header_row_set or (a is None and b is None and c is None and d is None):
            r += 1
            continue

        # A row with numeric data in C and D starts or continues a block
        if c is not None and d is not None:
            # --- start of a new block ---
            organ        = None
            groups       = []          # [{"name": str, "rows": [int]}]
            current_grp  = None
            hk_name      = None
            test_name    = None

            # Determine header row: the row just before this one
            # (either an existing header row or we'll insert one later)
            candidate_header = r - 1
            if candidate_header in header_row_set:
                header_row = candidate_header
                hk_name   = ws.cell(row=header_row, column=COL_HK).value
                test_name = ws.cell(row=header_row, column=COL_TEST).value
            else:
                # No pre-existing header — we'll insert one
                header_row = None   # signal to caller to insert

            # Scan data rows of this block
            rr = r
            while rr <= max_row:
                av = ws.cell(row=rr, column=COL_ORGAN).value
                bv = ws.cell(row=rr, column=COL_LABEL).value
                cv = ws.cell(row=rr, column=COL_HK).value
                dv = ws.cell(row=rr, column=COL_TEST).value

                # Blank row or another header row → end of block
                if (av is None and bv is None and cv is None and dv is None) or \
                   rr in header_row_set and rr != r:
                    break

                if av is not None:
                    organ = str(av).strip()

                if cv is not None:  # has HK data → data row
                    if is_continuation(bv):
                        if current_grp is not None:
                            current_grp["rows"].append(rr)
                    else:
                        current_grp = {"name": str(bv).strip(), "rows": [rr]}
                        groups.append(current_grp)

                rr += 1

            if organ and groups:
                ctrl_rows   = groups[0]["rows"]
                ctrl_label  = groups[0]["name"]
                drug_rows   = []
                drug_labels = []
                for g in groups[1:]:
                    drug_rows.extend(g["rows"])
                    drug_labels.append(g["name"])

                blocks.append({
                    "organ":       organ,
                    "header_row":  header_row,   # None → needs insertion
                    "ctrl_rows":   ctrl_rows,
                    "ctrl_label":  ctrl_label,
                    "drug_rows":   drug_rows,
                    "drug_labels": drug_labels,
                    "hk_name":     hk_name,
                    "test_name":   test_name,
                    "_block_start": r,           # used for insertion offset
                    "_block_end":   rr - 1,
                })

            r = rr  # jump past the block we just parsed
        else:
            r += 1

    return blocks


# ── Header insertion ──────────────────────────────────────────────────────────

def ensure_headers(ws, blocks, hk_col_name: str = "HK Gene",
                   test_col_name: str = "Test Gene") -> list[dict]:
    """
    For any block whose header_row is None, insert a header row above the
    first data row and adjust all subsequent row references accordingly.
    Returns the (mutated) blocks list.
    """
    insert_count = 0   # cumulative offset from prior insertions

    for blk in blocks:
        # Shift all row refs by insertions already done above this block
        blk["ctrl_rows"]  = [r + insert_count for r in blk["ctrl_rows"]]
        blk["drug_rows"]  = [r + insert_count for r in blk["drug_rows"]]
        blk["_block_start"] += insert_count
        blk["_block_end"]   += insert_count

        if blk["header_row"] is None:
            insert_at = blk["_block_start"]
            ws.insert_rows(insert_at)
            # Write static header text
            ws.cell(row=insert_at, column=COL_HK).value    = blk["hk_name"]   or hk_col_name
            ws.cell(row=insert_at, column=COL_TEST).value  = blk["test_name"] or test_col_name
            ws.cell(row=insert_at, column=COL_DELTA).value = "delta"
            ws.cell(row=insert_at, column=COL_REL).value   = "Relative Expr"

            blk["header_row"] = insert_at
            blk["ctrl_rows"]  = [r + 1 for r in blk["ctrl_rows"]]
            blk["drug_rows"]  = [r + 1 for r in blk["drug_rows"]]
            blk["_block_end"] += 1
            insert_count += 1

    return blocks


# ── Formula writer ────────────────────────────────────────────────────────────

def write_formulas(ws, blocks) -> None:
    """Write delta and relative-expression formulas into cols E and F."""
    for blk in blocks:
        h        = blk["header_row"]
        all_rows = blk["ctrl_rows"] + blk["drug_rows"]

        # E: delta = test gene − housekeeping gene
        for r in all_rows:
            ws.cell(row=r, column=COL_DELTA).value = (
                f"={cl(COL_TEST)}{r}-{cl(COL_HK)}{r}"
            )

        # F header row: average of control deltas
        ctrl_refs = ",".join(f"{cl(COL_DELTA)}{r}" for r in blk["ctrl_rows"])
        ws.cell(row=h, column=COL_REL).value = f"=AVERAGE({ctrl_refs})"

        # F data rows: 2^(avg_ctrl_delta − delta)
        for r in all_rows:
            ws.cell(row=r, column=COL_REL).value = (
                f"=2^({cl(COL_REL)}{h}-{cl(COL_DELTA)}{r})"
            )


# ── Styling ───────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _border() -> Border:
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

CENTER = Alignment(horizontal="center", vertical="center")

def style_workbook(ws, blocks) -> None:
    """Apply formatting to header and data rows of every block."""
    ws.column_dimensions[cl(COL_ORGAN)].width = 10
    ws.column_dimensions[cl(COL_LABEL)].width = 20
    ws.column_dimensions[cl(COL_HK)].width    = 14
    ws.column_dimensions[cl(COL_TEST)].width   = 14
    ws.column_dimensions[cl(COL_DELTA)].width  = 12
    ws.column_dimensions[cl(COL_REL)].width    = 18

    hdr_fill  = _fill(HDR_HEX)
    ctrl_fill = _fill(CTRL_HEX)
    drug_fill = _fill(DRUG_HEX)
    border    = _border()
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    std_font  = Font(name="Arial", size=10)

    for blk in blocks:
        h = blk["header_row"]

        # Header row
        for col in range(1, COL_REL + 1):
            cell = ws.cell(row=h, column=col)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = CENTER
            cell.border    = border

        # Control rows (gray)
        for r in blk["ctrl_rows"]:
            for col in range(1, COL_REL + 1):
                cell = ws.cell(row=r, column=col)
                cell.font      = std_font
                cell.fill      = ctrl_fill
                cell.alignment = CENTER
                cell.border    = border

        # Drug rows (red)
        for r in blk["drug_rows"]:
            for col in range(1, COL_REL + 1):
                cell = ws.cell(row=r, column=col)
                cell.font      = std_font
                cell.fill      = drug_fill
                cell.alignment = CENTER
                cell.border    = border


def add_legend_row(ws, blocks) -> list[dict]:
    """
    Prepend a single legend row at the very top of the sheet and shift
    all block row refs down by 1.
    """
    ws.insert_rows(1)
    ws["A1"] = "Legend:"
    ws["B1"] = "Gray = Control   |   Red = Drug-treated"
    ws["A1"].font = Font(bold=True, name="Arial", size=10)
    ws["B1"].font = Font(italic=True, name="Arial", size=10)

    for blk in blocks:
        blk["header_row"] += 1
        blk["ctrl_rows"]   = [r + 1 for r in blk["ctrl_rows"]]
        blk["drug_rows"]   = [r + 1 for r in blk["drug_rows"]]

    return blocks


# ── Chart data extractor ──────────────────────────────────────────────────────

def extract_chart_data(ws_data_only, blocks) -> list[dict]:
    """
    Read back calculated values from the recalculated workbook and return
    per-organ means for chart plotting.
    """
    results = []
    for blk in blocks:
        def _mean(rows):
            vals = []
            for r in rows:
                v = ws_data_only.cell(row=r, column=COL_REL).value
                if v is not None:
                    try:
                        vals.append(float(v))
                    except (TypeError, ValueError):
                        pass
            return float(np.mean(vals)) if vals else 0.0

        results.append({
            "organ":       blk["organ"],
            "ctrl_mean":   _mean(blk["ctrl_rows"]),
            "drug_mean":   _mean(blk["drug_rows"]),
            "ctrl_label":  blk["ctrl_label"],
            "drug_labels": blk["drug_labels"],
        })
    return results


# ── Chart renderer ────────────────────────────────────────────────────────────

def render_chart(chart_data: list[dict]) -> bytes:
    """
    Render the grouped bar chart and return raw PNG bytes (for Streamlit
    download or st.image display).
    """
    organs     = [d["organ"]     for d in chart_data]
    ctrl_means = [d["ctrl_mean"] for d in chart_data]
    drug_means = [d["drug_mean"] for d in chart_data]
    n          = len(organs)

    ctrl_label = chart_data[0]["ctrl_label"] if chart_data else "Control"
    drug_label = ", ".join(chart_data[0]["drug_labels"]) if chart_data else "Treatment"

    group_spacing = 1.4
    bar_width     = 0.22
    bar_gap       = 0.10
    half_span     = bar_width / 2 + bar_gap / 2
    x             = np.arange(n) * group_spacing

    fig, ax = plt.subplots(figsize=(max(5, 2.4 * n + 1.8), 5.5))
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    ax.bar(x - half_span, ctrl_means, width=bar_width, color=CTRL_MPL, zorder=3)
    ax.bar(x + half_span, drug_means, width=bar_width, color=DRUG_MPL, zorder=3)

    ax.set_xticks(x)
    ax.set_xticklabels(organs, fontsize=13)
    ax.set_ylabel("Relative Expression", fontsize=12)
    ax.set_ylim(bottom=0)

    ax.yaxis.grid(True, color="#E8E8E8", linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#AAAAAA")
    ax.spines["bottom"].set_color("#AAAAAA")
    ax.tick_params(axis="both", which="both", length=0, labelsize=11)
    ax.set_xlim(
        -group_spacing * 0.55,
        (n - 1) * group_spacing + group_spacing * 0.55
    )

    ctrl_patch = mpatches.Patch(color=CTRL_MPL, label=ctrl_label)
    drug_patch = mpatches.Patch(color=DRUG_MPL, label=drug_label)
    ax.legend(
        handles=[ctrl_patch, drug_patch],
        loc="upper center",
        bbox_to_anchor=(0.5, 1.13),
        ncol=2,
        frameon=False,
        fontsize=11,
        handlelength=1.0,
        handleheight=1.0,
    )

    plt.tight_layout(rect=[0, 0, 1, 0.95])

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=180, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close()
    buf.seek(0)
    return buf.read()


# ── Top-level pipeline (called by Streamlit app) ──────────────────────────────

def run_pipeline(file_bytes: bytes) -> dict:
    """
    Accept raw .xlsx bytes, run the full analysis, and return:
    {
        "xlsx_bytes"  : bytes   — completed workbook
        "chart_bytes" : bytes   — PNG chart
        "chart_data"  : list    — per-organ means (for optional table display)
        "blocks"      : list    — parsed block metadata
        "errors"      : str     — non-empty if something went wrong
    }
    """
    import subprocess, json, tempfile, os, shutil

    result = {"xlsx_bytes": None, "chart_bytes": None,
              "chart_data": [], "blocks": [], "errors": ""}

    # --- 1. Load ---
    try:
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
    except Exception as e:
        result["errors"] = f"Could not open file: {e}"
        return result

    # --- 2. Parse ---
    blocks = parse_blocks(ws)
    if not blocks:
        result["errors"] = (
            "No data blocks detected. Make sure columns A–D are filled in "
            "and each organ group is separated by a blank row."
        )
        return result
    result["blocks"] = blocks

    # --- 3. Ensure header rows exist (insert if needed) ---
    blocks = ensure_headers(ws, blocks)

    # --- 4. Write formulas ---
    write_formulas(ws, blocks)

    # --- 5. Style ---
    style_workbook(ws, blocks)

    # --- 6. Add legend row at top ---
    blocks = add_legend_row(ws, blocks)

    # Re-write formulas after row shifts from legend insertion
    write_formulas(ws, blocks)

    # --- 7. Save to temp file for recalculation ---
    tmp_dir = tempfile.mkdtemp()
    tmp_path = os.path.join(tmp_dir, "output.xlsx")
    try:
        wb.save(tmp_path)

        # --- 8. Recalculate with LibreOffice ---
        recalc = subprocess.run(
            ["python3", "/mnt/skills/public/xlsx/scripts/recalc.py", tmp_path],
            capture_output=True, text=True, timeout=60
        )
        try:
            info = json.loads(recalc.stdout)
            if info.get("status") == "errors_found":
                result["errors"] = (
                    f"Formula errors after calculation: {info.get('error_summary')}"
                )
                return result
        except Exception:
            # recalc script not available (e.g. local dev) — continue with unrecalculated
            pass

        # --- 9. Read back calculated values ---
        wb2 = load_workbook(tmp_path, data_only=True)
        ws2 = wb2.active
        chart_data = extract_chart_data(ws2, blocks)
        result["chart_data"] = chart_data

        # --- 10. Read final xlsx bytes ---
        with open(tmp_path, "rb") as f:
            result["xlsx_bytes"] = f.read()

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    # --- 11. Render chart ---
    try:
        result["chart_bytes"] = render_chart(chart_data)
    except Exception as e:
        result["errors"] = f"Chart rendering failed: {e}"

    return result
